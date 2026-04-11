import re
from datetime import datetime


REQUIRED_COLUMNS = [
    "first_name",
    "last_name",
    "date_of_birth",
    "gender",
    "admission_date",
]

OPTIONAL_COLUMNS = [
    "other_names",
    "address",
    "previous_school",
    "class_id",
    "academic_year_id",
]

VALID_GENDERS = ["male", "female", "other"]
VALID_RELATIONSHIPS = ["father", "mother", "guardian", "sibling", "other"]

# Matches guardian_1_first_name, guardian_2_phone etc.
GUARDIAN_COLUMN_PATTERN = re.compile(
    r"^guardian_(\d+)_(.+)$"
)


def parse_students_from_excel(file):
    """
    Parses an uploaded Excel file into a list of student dicts.

    Returns:
        students  : list of validated student dicts
        row_errors: list of { row, errors } for invalid rows
    """
    import openpyxl

    try:
        workbook = openpyxl.load_workbook(file, read_only=True, data_only=True)
    except Exception:
        raise ValueError(
            "Could not read the file. "
            "Please upload a valid .xlsx Excel file."
        )

    sheet = workbook.active

    # ── Extract headers from first row ────────────────────────────
    headers = []
    for cell in next(sheet.iter_rows(min_row=1, max_row=1)):
        value = str(cell.value).strip().lower() if cell.value else ""
        headers.append(value)

    if not headers:
        raise ValueError("The Excel file appears to be empty.")

    # ── Check required columns exist ──────────────────────────────
    missing = [col for col in REQUIRED_COLUMNS if col not in headers]
    if missing:
        raise ValueError(
            f"Missing required column(s): {', '.join(missing)}. "
            f"Please use the provided template."
        )

    students = []
    row_errors = []

    for row_index, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        # Skip completely empty rows
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        row_dict = {
            headers[i]: (
                str(row[i]).strip()
                if row[i] is not None else ""
            )
            for i in range(min(len(headers), len(row)))
        }

        errors = []
        student = {}
        guardians = {}

        # ── Parse each cell ───────────────────────────────────────
        for key, value in row_dict.items():

            # Guardian columns
            guardian_match = GUARDIAN_COLUMN_PATTERN.match(key)
            if guardian_match:
                g_index = guardian_match.group(1)
                g_field = guardian_match.group(2)
                if g_index not in guardians:
                    guardians[g_index] = {}
                guardians[g_index][g_field] = value
                continue

            # Student columns
            student[key] = value

        # ── Validate required fields ──────────────────────────────
        for col in REQUIRED_COLUMNS:
            if not student.get(col):
                errors.append(f"'{col}' is required.")

        # ── Validate gender ───────────────────────────────────────
        gender = student.get("gender", "").lower()
        if gender and gender not in VALID_GENDERS:
            errors.append(
                f"Invalid gender '{gender}'. "
                f"Must be one of: {', '.join(VALID_GENDERS)}."
            )
        else:
            student["gender"] = gender

        # ── Validate and parse dates ──────────────────────────────
        for date_field in ["date_of_birth", "admission_date"]:
            raw = student.get(date_field, "")
            parsed = _parse_date(raw)
            if raw and parsed is None:
                errors.append(
                    f"Invalid date format for '{date_field}': '{raw}'. "
                    f"Use YYYY-MM-DD."
                )
            else:
                student[date_field] = parsed

        # ── Clean optional fields ─────────────────────────────────
        student.setdefault("other_names", "")
        student.setdefault("address", "")
        student.setdefault("previous_school", "")

        # Strip empty UUIDs
        for uuid_field in ["class_id", "academic_year_id"]:
            val = student.get(uuid_field, "").strip()
            student[uuid_field] = val if val else None

        # ── Validate guardians ────────────────────────────────────
        parsed_guardians = []
        for g_index, g_data in guardians.items():
            g_errors, g_parsed = _validate_guardian(
                g_data, g_index, row_index
            )
            errors.extend(g_errors)
            if g_parsed:
                parsed_guardians.append(g_parsed)

        student["guardians"] = parsed_guardians

        # ── Collect row result ────────────────────────────────────
        if errors:
            row_errors.append({
                "row": row_index,
                "name": (
                    f"{student.get('first_name', '')} "
                    f"{student.get('last_name', '')}"
                ).strip() or f"Row {row_index}",
                "errors": errors,
            })
        else:
            students.append(student)

    workbook.close()
    return students, row_errors


def _parse_date(value):
    """Tries multiple date formats. Returns date object or None."""
    if not value:
        return None

    # Already a date/datetime object (openpyxl can return these)
    if hasattr(value, "date"):
        return value.date()
    if hasattr(value, "year"):
        return value

    formats = ["%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"]
    for fmt in formats:
        try:
            return datetime.strptime(str(value).strip(), fmt).date()
        except ValueError:
            continue
    return None


def _validate_guardian(g_data, g_index, row_index):
    """Validates a single guardian dict. Returns (errors, parsed)."""
    errors = []

    first_name = g_data.get("first_name", "").strip()
    last_name = g_data.get("last_name", "").strip()
    phone = g_data.get("phone", "").strip()
    relationship = g_data.get("relationship", "").strip().lower()
    is_primary_raw = g_data.get("is_primary", "false")

    # Skip guardian if all fields are empty
    if not any([first_name, last_name, phone, relationship]):
        return [], None

    if not first_name:
        errors.append(f"Guardian {g_index}: 'first_name' is required.")
    if not last_name:
        errors.append(f"Guardian {g_index}: 'last_name' is required.")
    if not phone:
        errors.append(f"Guardian {g_index}: 'phone' is required.")
    if not relationship:
        errors.append(f"Guardian {g_index}: 'relationship' is required.")
    elif relationship not in VALID_RELATIONSHIPS:
        errors.append(
            f"Guardian {g_index}: invalid relationship '{relationship}'. "
            f"Must be one of: {', '.join(VALID_RELATIONSHIPS)}."
        )

    if errors:
        return errors, None

    # Parse is_primary
    is_primary = str(is_primary_raw).strip().lower() in (
        "true", "1", "yes"
    )

    return [], {
        "first_name": first_name,
        "last_name": last_name,
        "relationship": relationship,
        "phone": phone,
        "email": g_data.get("email", "").strip(),
        "address": g_data.get("address", "").strip(),
        "is_primary": is_primary,
    }