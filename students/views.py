from rest_framework import generics
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from rest_framework.permissions import AllowAny

from academics.models import AcademicYear
from core.permissions import IsAdmin, IsAdminOrTeacher
from core.responses import ApiResponse
from core.mixins import AuditLogMixin, ExportMixin
from core.models import AuditLog
from core.utils import log_action
from students.utils import parse_students_from_excel
# from subscriptions.utils import check_limit

from .models import Student, Guardian, Enrollment
from .serializers import (
    StudentListSerializer, StudentMinimalSerializer, StudentSerializer, StudentCreateSerializer,
    GuardianSerializer, EnrollStudentSerializer,
    EnrollmentSerializer, StudentUpdateSerializer,
)
from .filters import StudentFilter
from django.db import transaction


# ─── Students ────────────────────────────────────────────────────

class StudentBulkCreateView(APIView):
    """
    Bulk create students from an uploaded Excel file.

    Payload: multipart/form-data
        file = .xlsx file (required)
        class_id = UUID (optional — applied to ALL students if provided)
        academic_year_id = UUID (optional — applied to ALL students if provided)

    Individual students in the Excel can override class_id
    and academic_year_id per row.
    """
    permission_classes = [IsAdmin]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        school = request.user.school

        # ── Validate file ─────────────────────────────────────────
        file = request.FILES.get("file")
        if not file:
            return ApiResponse.error(
                message="No file uploaded. Please upload an .xlsx file.",
                status_code=400,
            )

        if not file.name.endswith(".xlsx"):
            return ApiResponse.error(
                message="Invalid file type. Only .xlsx files are supported.",
                status_code=400,
            )

        # ── Global fallback class/year ────────────────────────────
        # Applied to any student row that doesn't have its own
        global_class_id = request.data.get("class_id") or None
        if global_class_id:
            global_academic_year_id = AcademicYear.objects.filter(
                is_current=True, school=school
            ).first().id
        else:
            global_academic_year_id = None

        # ── Parse Excel ───────────────────────────────────────────
        try:
            students_data, parse_errors = parse_students_from_excel(file)
        except ValueError as e:
            return ApiResponse.error(message=str(e), status_code=400)

        if not students_data and parse_errors:
            return ApiResponse.error(
                message="All rows failed validation. No students were created.",
                data={
                    "summary": {
                        "total_submitted": len(parse_errors),
                        "created": 0,
                        "failed": len(parse_errors),
                        "parse_errors": parse_errors,
                    }
                },
                status_code=400,
            )

        # ── Plan limit check ──────────────────────────────────────
        from subscriptions.utils import get_active_subscription
        subscription = get_active_subscription(school)
        if subscription and subscription.plan.max_students is not None:
            current_count = school.students.filter(
                status="active"
            ).count()
            remaining = subscription.plan.max_students - current_count
            if len(students_data) > remaining:
                return ApiResponse.error(
                    message=(
                        f"This import would exceed your plan's student limit. "
                        f"You can add {remaining} more student(s). "
                        f"This file contains {len(students_data)} valid records."
                    ),
                    status_code=400,
                )

        # ── Create students ───────────────────────────────────────
        created = []
        failed = [*parse_errors]  # start with parse errors

        for student_data in students_data:
            row_num = student_data.pop("_row", None)

            # Apply global fallbacks if row doesn't have its own
            if not student_data.get("class_id"):
                student_data["class_id"] = global_class_id
            if not student_data.get("academic_year_id"):
                student_data["academic_year_id"] = global_academic_year_id

            try:
                with transaction.atomic():
                    student = self._create_student(
                        student_data=student_data,
                        school=school,
                    )
                    created.append({
                        "student_id": str(student.id),
                        "student_id_number": student.student_id,
                        "name": student.full_name,
                    })
            except Exception as e:
                failed.append({
                    "row": row_num,
                    "name": (
                        f"{student_data.get('first_name', '')} "
                        f"{student_data.get('last_name', '')}"
                    ).strip(),
                    "errors": [str(e)],
                })

        # ── Audit log ─────────────────────────────────────────────
        if created:
            log_action(
                action=AuditLog.Action.CREATE,
                resource="Student",
                description=(
                    f"Bulk Excel import: {len(created)} created, "
                    f"{len(failed)} failed"
                ),
                request=request,
                metadata={
                    "created_count": len(created),
                    "failed_count": len(failed),
                    "file_name": file.name,
                },
            )

        total_submitted = len(students_data) + len(parse_errors)

        return ApiResponse.success(
            data={
                "summary": {
                    "total_submitted": total_submitted,
                    "created": len(created),
                    "failed": len(failed),
                    "success_rate": (
                        f"{round((len(created) / total_submitted) * 100)}%"
                        if total_submitted > 0 else "0%"
                    ),
                },
                "created": created,
                "failed": failed,
            },
            message=(
                f"{len(created)} of {total_submitted} students "
                f"imported successfully."
                + (
                    f" {len(failed)} failed — check 'failed' for details."
                    if failed else ""
                )
            ),
        )

    @staticmethod
    def _create_student(student_data, school):
        from academics.models import Class, AcademicYear

        guardians_data = student_data.pop("guardians", [])
        class_id = student_data.pop("class_id", None)
        academic_year_id = student_data.pop("academic_year_id", None)

        student = Student.objects.create(
            school=school, **student_data
        )

        for guardian_data in guardians_data:
            Guardian.objects.create(
                student=student,
                school=school,
                **guardian_data,
            )

        if class_id and academic_year_id:
            try:
                klass = Class.objects.get(id=class_id, school=school)
                academic_year = AcademicYear.objects.get(
                    id=academic_year_id, school=school
                )
                Enrollment.objects.create(
                    school=school,
                    student=student,
                    klass=klass,
                    academic_year=academic_year,
                )
            except (Class.DoesNotExist, AcademicYear.DoesNotExist):
                pass

        return student

class ClassBulkStudentsView(APIView):
    """
    Bulk create students for a specific class from an Excel file.
    Each student is enrolled in the given class with one guardian.

    POST /api/v1/students/class-bulk/
        class_id (required) — UUID of the class to enroll in
        file (required) — .xlsx file
        photos (optional) — .zip file with student photos named by photo_filename column

    GET  /api/v1/students/class-bulk/ — downloads the template
    """
    permission_classes = [IsAdmin]
    parser_classes = [MultiPartParser, FormParser]

    def get(self, request):
        return self._generate_template()

    def post(self, request):
        school = request.user.school
        class_id = request.data.get("class_id")
        if not class_id:
            return ApiResponse.error(
                message="class_id is required.",
                status_code=400,
            )

        # ── Validate class exists and is active ──────────────────────
        from academics.models import Class, AcademicYear
        try:
            print(class_id)
            klass = Class.objects.get(id=class_id, school=school, is_active=True)
        except Class.DoesNotExist:
            return ApiResponse.error(
                message="Class not found or inactive.",
                status_code=404,
            )

        # ── Resolve current academic year ────────────────────────────
        current_term = school.terms.filter(is_current=True).first()
        if not current_term:
            return ApiResponse.error(
                message="No active academic year. Please set up an academic year and term first.",
                status_code=400,
            )
        academic_year = current_term.academic_year

        # ── Validate file ───────────────────────────────────────────
        file = request.FILES.get("file")
        if not file:
            return ApiResponse.error(
                message="No file uploaded. Please upload an .xlsx file.",
                status_code=400,
            )
        if not file.name.endswith(".xlsx"):
            return ApiResponse.error(
                message="Invalid file type. Only .xlsx files are supported.",
                status_code=400,
            )

        # ── Parse Excel ─────────────────────────────────────────────
        try:
            students_data, parse_errors = self._parse_class_excel(file)
        except ValueError as e:
            return ApiResponse.error(message=str(e), status_code=400)

        if not students_data and parse_errors:
            return ApiResponse.error(
                message="All rows failed validation.",
                data={
                    "summary": {
                        "total_submitted": len(parse_errors),
                        "created": 0,
                        "failed": len(parse_errors),
                        "parse_errors": parse_errors,
                    }
                },
                status_code=400,
            )

        
        # ── Extract photos ZIP if provided ──────────────────────────
        import zipfile
        import tempfile
        import os

        photos_zip = request.FILES.get("photos")
        photo_map = {}
        temp_dir = None
        if photos_zip:
            if not photos_zip.name.endswith(".zip"):
                return ApiResponse.error(
                    message="Photos file must be a .zip archive.",
                    status_code=400,
                )
            temp_dir = tempfile.mkdtemp()
            try:
                with zipfile.ZipFile(photos_zip) as zf:
                    zf.extractall(temp_dir)
                    for root, dirs, files in os.walk(temp_dir):
                        for fname in files:
                            photo_map[fname.lower()] = os.path.join(root, fname)
            except zipfile.BadZipFile:
                import shutil
                shutil.rmtree(temp_dir, ignore_errors=True)
                return ApiResponse.error(
                    message="Invalid ZIP file for photos.",
                    status_code=400,
                )

        # ── Create students ─────────────────────────────────────────
        created = []
        failed = [*parse_errors]

        for student_data in students_data:
            row_num = student_data.pop("_row", None)
            photo_filename = student_data.pop("photo_filename", None)

            try:
                with transaction.atomic():
                    # Pop guardian before creating student
                    guardian_data = student_data.pop("guardian", None)

                    student = Student.objects.create(
                        school=school, **student_data
                    )

                    # Create guardian
                    if guardian_data:
                        Guardian.objects.create(
                            student=student,
                            school=school,
                            **guardian_data,
                        )

                    # Enroll in class
                    Enrollment.objects.create(
                        school=school,
                        student=student,
                        klass=klass,
                        academic_year=academic_year,
                    )

                    # Handle profile photo from ZIP
                    if photo_filename and photo_map:
                        normalized = photo_filename.strip().lower()
                        photo_path = photo_map.get(normalized)
                        if photo_path:
                            from django.core.files import File
                            with open(photo_path, "rb") as f:
                                student.profile_photo.save(
                                    normalized, File(f), save=True
                                )

                    created.append({
                        "student_id": str(student.id),
                        "student_id_number": student.student_id,
                        "name": student.full_name,
                    })
            except Exception as e:
                failed.append({
                    "row": row_num,
                    "name": (
                        f"{student_data.get('first_name', '')} "
                        f"{student_data.get('last_name', '')}"
                    ).strip(),
                    "errors": [str(e)],
                })

        # ── Cleanup temp photos ─────────────────────────────────────
        if temp_dir:
            import shutil
            shutil.rmtree(temp_dir, ignore_errors=True)

        # ── Audit log ───────────────────────────────────────────────
        if created:
            log_action(
                action=AuditLog.Action.CREATE,
                resource="Student",
                description=(
                    f"Class bulk import: {len(created)} created in "
                    f"{klass.name}, {len(failed)} failed"
                ),
                request=request,
                metadata={
                    "created_count": len(created),
                    "failed_count": len(failed),
                    "class_id": str(class_id),
                    "class_name": klass.name,
                    "academic_year": academic_year.name,
                    "file_name": file.name,
                    "photos_included": bool(photos_zip),
                },
            )

        total_submitted = len(students_data) + len(parse_errors)

        return ApiResponse.success(
            data={
                "summary": {
                    "class_name": klass.name,
                    "academic_year": academic_year.name,
                    "total_submitted": total_submitted,
                    "created": len(created),
                    "failed": len(failed),
                    "success_rate": (
                        f"{round((len(created) / total_submitted) * 100)}%"
                        if total_submitted > 0 else "0%"
                    ),
                },
                "created": created,
                "failed": failed,
            },
            message=(
                f"{len(created)} of {total_submitted} students imported "
                f"into {klass.name} successfully."
                + (
                    f" {len(failed)} failed — check 'failed' for details."
                    if failed else ""
                )
            ),
        )

    # ── Parsing ─────────────────────────────────────────────────────

    @staticmethod
    def _parse_class_excel(file):
        """
        Parses the simplified class-bulk Excel file.
        Returns (students_data, row_errors).
        """
        import openpyxl

        try:
            workbook = openpyxl.load_workbook(file, read_only=True, data_only=True)
        except Exception:
            raise ValueError("Could not read the file. Please upload a valid .xlsx Excel file.")

        sheet = workbook.active

        headers = []
        for cell in next(sheet.iter_rows(min_row=1, max_row=1)):
            value = str(cell.value).strip().lower() if cell.value else ""
            headers.append(value)

        if not headers:
            raise ValueError("The Excel file appears to be empty.")

        required = ["first_name", "last_name", "date_of_birth", "gender"]
        missing = [col for col in required if col not in headers]
        if missing:
            raise ValueError(
                f"Missing required column(s): {', '.join(missing)}. "
                f"Please use the provided template."
            )

        valid_genders = ["male", "female", "other"]
        valid_relationships = ["father", "mother", "guardian", "sibling", "other"]

        students = []
        row_errors = []

        for row_index, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            row_dict = {
                headers[i]: (
                    str(row[i]).strip() if row[i] is not None else ""
                )
                for i in range(min(len(headers), len(row)))
            }

            errors = []
            student = {}

            for key, value in row_dict.items():
                student[key] = value

            # Validate required fields
            for col in required:
                if not student.get(col):
                    errors.append(f"'{col}' is required.")

            # Validate gender
            gender = student.get("gender", "").lower()
            if gender and gender not in valid_genders:
                errors.append(
                    f"Invalid gender '{gender}'. "
                    f"Must be one of: {', '.join(valid_genders)}."
                )
            else:
                student["gender"] = gender

            # Parse dates
            def parse_date(raw):
                if not raw:
                    return None
                from datetime import datetime as dt
                formats = ["%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"]
                for fmt in formats:
                    try:
                        return dt.strptime(str(raw).strip(), fmt).date()
                    except ValueError:
                        continue
                return None

            for date_field in ["date_of_birth", "admission_date"]:
                raw = student.get(date_field, "")
                parsed = parse_date(raw)
                if raw and parsed is None:
                    errors.append(f"Invalid date format for '{date_field}': '{raw}'. Use YYYY-MM-DD.")
                else:
                    student[date_field] = parsed

            # Clean optional fields
            student.setdefault("other_names", "")
            student.setdefault("address", "")
            student.setdefault("previous_school", "")
            student.setdefault("email", "")
            student.setdefault("phone_number", "")

            # Guardian (single)
            guardian = {}
            g_first_name = student.pop("guardian_first_name", "").strip()
            g_last_name = student.pop("guardian_last_name", "").strip()
            g_relationship = student.pop("guardian_relationship", "").strip().lower()
            g_phone = student.pop("guardian_phone", "").strip()
            g_email = student.pop("guardian_email", "").strip()
            g_address = student.pop("guardian_address", "").strip()
            g_is_primary = student.pop("guardian_is_primary", "").strip().lower() in ("true", "1", "yes")

            has_guardian = any([g_first_name, g_last_name, g_relationship, g_phone])
            if has_guardian:
                if not g_first_name:
                    errors.append("Guardian 'first_name' is required.")
                if not g_last_name:
                    errors.append("Guardian 'last_name' is required.")
                if not g_relationship:
                    errors.append("Guardian 'relationship' is required.")
                elif g_relationship not in valid_relationships:
                    errors.append(
                        f"Invalid guardian relationship '{g_relationship}'. "
                        f"Must be one of: {', '.join(valid_relationships)}."
                    )
                if not g_phone:
                    errors.append("Guardian 'phone' is required.")

                if not errors:
                    guardian = {
                        "first_name": g_first_name,
                        "last_name": g_last_name,
                        "relationship": g_relationship,
                        "phone": g_phone,
                        "email": g_email,
                        "address": g_address,
                        "is_primary": g_is_primary,
                    }

            student["guardian"] = guardian if guardian else None

            # Photo filename
            photo_filename = student.pop("photo_filename", "").strip()
            student["photo_filename"] = photo_filename if photo_filename else None

            if errors:
                row_errors.append({
                    "row": row_index,
                    "name": (
                        f"{student.get('first_name', '')} "
                        f"{student.get('last_name', '')}"
                    ).strip() or f"Row {row_index}",
                    "errors": errors,
                })
            else:
                student["_row"] = row_index
                students.append(student)

        workbook.close()
        return students, row_errors

    # ── Template generation ────────────────────────────────────────

    @staticmethod
    def _generate_template():
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Students"

        headers = [
            "first_name",
            "last_name",
            "other_names",
            "date_of_birth",
            "gender",
            "admission_date",
            "address",
            "previous_school",
            "email",
            "phone_number",
            "photo_filename",
            "guardian_first_name",
            "guardian_last_name",
            "guardian_relationship",
            "guardian_phone",
            "guardian_email",
            "guardian_address",
            "guardian_is_primary",
        ]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            sheet.column_dimensions[cell.column_letter].width = max(len(header) + 4, 18)

        sample = [
            "Ekow", "Mensah", "Kofi",
            "2010-05-14", "male", "2025-09-01",
            "123 Adum Street, Kumasi", "Good Shepherd Academy",
            "ekow.mensah@school.com", "0244000001",
            "ekow_photo.jpg",
            "Kwame", "Mensah", "father",
            "0244000002", "kwame@email.com", "", "true",
        ]
        for col_num, value in enumerate(sample, start=1):
            cell = sheet.cell(row=2, column=col_num, value=value)
            cell.alignment = Alignment(horizontal="left")

        # Notes sheet
        notes_sheet = workbook.create_sheet(title="Notes")
        notes = [
            ["Field", "Required", "Format / Allowed Values"],
            ["first_name", "Yes", "Text"],
            ["last_name", "Yes", "Text"],
            ["other_names", "No", "Text"],
            ["date_of_birth", "Yes", "YYYY-MM-DD or DD/MM/YYYY"],
            ["gender", "Yes", "male / female / other"],
            ["admission_date", "No", "YYYY-MM-DD or DD/MM/YYYY"],
            ["address", "No", "Text"],
            ["previous_school", "No", "Text"],
            ["email", "No", "Email address"],
            ["phone_number", "No", "Phone number"],
            ["photo_filename", "No", "Filename in the uploaded photos ZIP (e.g., student1.jpg)"],
            ["guardian_first_name", "No", "Text"],
            ["guardian_last_name", "No", "Text"],
            ["guardian_relationship", "No", "father / mother / guardian / sibling / other"],
            ["guardian_phone", "No", "Phone number"],
            ["guardian_email", "No", "Email address"],
            ["guardian_address", "No", "Text"],
            ["guardian_is_primary", "No", "true / false"],
        ]
        notes_header_font = Font(bold=True)
        for row_index, row_data in enumerate(notes, start=1):
            for col_index, value in enumerate(row_data, start=1):
                cell = notes_sheet.cell(row=row_index, column=col_index, value=value)
                if row_index == 1:
                    cell.font = notes_header_font
                notes_sheet.column_dimensions[cell.column_letter].width = 40

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="skuumate_class_bulk_template.xlsx"'
        workbook.save(response)
        return response


class StudentExcelTemplateView(APIView):
    """
    Returns a downloadable Excel template for bulk student import.
    """
    permission_classes = [AllowAny]

    def get(self, request):
        import openpyxl
        from django.http import HttpResponse

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Students"

        headers = [
            "first_name",
            "last_name",
            "other_names",
            "date_of_birth",
            "gender",
            "admission_date",
            "address",
            "previous_school",
            "class_id",
            "academic_year_id",
            "guardian_1_first_name",
            "guardian_1_last_name",
            "guardian_1_relationship",
            "guardian_1_phone",
            "guardian_1_email",
            "guardian_1_is_primary",
            "guardian_2_first_name",
            "guardian_2_last_name",
            "guardian_2_relationship",
            "guardian_2_phone",
            "guardian_2_email",
            "guardian_2_is_primary",
        ]

        # ── Style the header row ──────────────────────────────────
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="2563EB",
            end_color="2563EB",
            fill_type="solid",
        )

        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            sheet.column_dimensions[
                cell.column_letter
            ].width = max(len(header) + 4, 16)

        # ── Add a sample row ──────────────────────────────────────
        sample = [
            "Ekow", "Mensah", "Kofi",
            "2010-05-14", "male", "2025-09-01",
            "123 Adum Street, Kumasi", "Good Shepherd Academy",
            "", "",   # class_id and academic_year_id left blank
            "Kwame", "Mensah", "father",
            "0244000001", "kwame@email.com", "true",
            "Akua", "Mensah", "mother",
            "0244000002", "", "false",
        ]

        for col_num, value in enumerate(sample, start=1):
            cell = sheet.cell(row=2, column=col_num, value=value)
            cell.alignment = Alignment(horizontal="left")

        # ── Add a notes sheet ─────────────────────────────────────
        notes_sheet = workbook.create_sheet(title="Notes")
        notes = [
            ["Field", "Required", "Format / Allowed Values"],
            ["first_name", "Yes", "Text"],
            ["last_name", "Yes", "Text"],
            ["other_names", "No", "Text"],
            ["date_of_birth", "Yes", "YYYY-MM-DD or DD/MM/YYYY"],
            ["gender", "Yes", "male / female / other"],
            ["admission_date", "Yes", "YYYY-MM-DD or DD/MM/YYYY"],
            ["address", "No", "Text"],
            ["previous_school", "No", "Text"],
            ["class_id", "No", "UUID from the system"],
            ["academic_year_id", "No", "UUID from the system"],
            ["guardian_X_first_name", "No", "Text"],
            ["guardian_X_last_name", "No", "Text"],
            ["guardian_X_relationship", "No", "father/mother/guardian/sibling/other"],
            ["guardian_X_phone", "No", "Phone number"],
            ["guardian_X_email", "No", "Email address"],
            ["guardian_X_is_primary", "No", "true / false"],
        ]

        notes_header_font = Font(bold=True)
        for row_index, row_data in enumerate(notes, start=1):
            for col_index, value in enumerate(row_data, start=1):
                cell = notes_sheet.cell(
                    row=row_index, column=col_index, value=value
                )
                if row_index == 1:
                    cell.font = notes_header_font
                notes_sheet.column_dimensions[
                    cell.column_letter
                ].width = 35

        # ── Return as downloadable file ───────────────────────────
        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.sheet"
            )
        )
        response["Content-Disposition"] = (
            'attachment; filename="skuumate_students_template.xlsx"'
        )
        workbook.save(response)
        return response

class StudentListCreateView(AuditLogMixin, ExportMixin, generics.ListCreateAPIView):
    permission_classes = [IsAdminOrTeacher]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_class = StudentFilter
    search_fields = [
        "first_name", "last_name",
        "other_names", "student_id",
    ]
    ordering_fields = [
        "first_name", "last_name",
        "admission_date", "created_at",
    ]
    ordering = ["last_name", "first_name"]
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    audit_resource = "Student"

    def get_permissions(self):
        if self.request.method == "POST":
            return [IsAdmin()]
        return [IsAdminOrTeacher()]

    def get_serializer_class(self):
        if self.request.method == "POST":
            return StudentCreateSerializer
        return StudentListSerializer

    def get_queryset(self):
        return Student.objects.filter(
            school=self.request.user.school
        ).prefetch_related("guardians", "enrollments__klass", "enrollments__academic_year")

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx["school"] = self.request.user.school
        return ctx

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = StudentListSerializer(page, many=True, context={"request": request})
            return self.get_paginated_response(serializer.data)
        serializer = StudentListSerializer(queryset, many=True, context={"request": request})
        return ApiResponse.success(data=serializer.data)

    def create(self, request, *args, **kwargs):
        # check_limit(request.user.school, "students") 
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        student = serializer.save()
        log_action(
            action=AuditLog.Action.CREATE,
            resource="Student",
            resource_id=str(student.pk),
            description=f"Student {student.full_name} registered",
            request=request,
        )
        return ApiResponse.created(
            data=StudentListSerializer(student).data,
            message="Student registered successfully.",
        )


class StudentDetailView(AuditLogMixin, generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAdminOrTeacher]
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    audit_resource = "Student"

    def get_permissions(self):
        if self.request.method in ["PUT","PATCH", "DELETE"]:
            return [IsAdmin()]
        return [IsAdminOrTeacher()]

    def get_serializer_class(self):
        if self.request.method in ["PATCH","PUT"]:
            return StudentUpdateSerializer
        return StudentSerializer

    def get_queryset(self):
        return Student.objects.filter(
            school=self.request.user.school
        ).prefetch_related("guardians", "enrollments__klass", "enrollments__academic_year")

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx["school"] = self.request.user.school
        return ctx

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        return ApiResponse.success(data=StudentSerializer(instance).data)

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = StudentUpdateSerializer(
            instance, data=request.data, partial=True,
            context={"school": request.user.school, "request":request}
        )
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return ApiResponse.success(
            data=StudentListSerializer(serializer.instance, context = {"request":request}).data,
            message="Student updated successfully.",
        )


class StudentExportView(ExportMixin, generics.ListAPIView):
    permission_classes = [IsAdminOrTeacher]
    serializer_class = StudentSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_class = StudentFilter
    search_fields = ["first_name", "last_name", "other_names", "student_id"]
    ordering = ["last_name", "first_name"]

    def get_queryset(self):
        return Student.objects.filter(
            school=self.request.user.school
        ).prefetch_related("guardians", "enrollments__klass", "enrollments__academic_year")

    def get(self, request, *args, **kwargs):
        return self.export(request, *args, **kwargs)


# ─── Guardians ────────────────────────────────────────────────────

class GuardianListCreateView(generics.ListCreateAPIView):
    permission_classes = [IsAdmin]
    serializer_class = GuardianSerializer

    def get_student(self, request, pk):
        try:
            return Student.objects.get(pk=pk, school=request.user.school)
        except Student.DoesNotExist:
            return None

    def get_queryset(self):
        student = self.get_student(self.request, self.kwargs["pk"])
        if not student:
            return Guardian.objects.none()
        return Guardian.objects.filter(student=student)

    def list(self, request, *args, **kwargs):
        student = self.get_student(request, kwargs["pk"])
        if not student:
            return ApiResponse.error(message="Student not found.", status_code=404)
        serializer = self.get_serializer(self.get_queryset(), many=True)
        return ApiResponse.success(data=serializer.data)

    def create(self, request, *args, **kwargs):
        student = self.get_student(request, kwargs["pk"])
        if not student:
            return ApiResponse.error(message="Student not found.", status_code=404)
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        guardian = serializer.save(student=student, school=request.user.school)
        log_action(
            action=AuditLog.Action.CREATE,
            resource="Guardian",
            resource_id=str(guardian.pk),
            description=f"Guardian {guardian} added to {student.full_name}",
            request=request,
        )
        return ApiResponse.created(
            data=serializer.data,
            message="Guardian added successfully.",
        )


class GuardianDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAdmin]
    serializer_class = GuardianSerializer

    def get_queryset(self):
        return Guardian.objects.filter(school=self.request.user.school)

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        return ApiResponse.success(data=self.get_serializer(instance).data)

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return ApiResponse.success(
            data=serializer.data,
            message="Guardian updated successfully.",
        )

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return ApiResponse.success(message="Guardian removed successfully.")


# ─── Enrollment ───────────────────────────────────────────────────

class StudentEnrollView(APIView):
    """Enroll a student into a class for a given academic year."""
    permission_classes = [IsAdmin]

    def get_student(self, request, pk):
        try:
            return Student.objects.get(pk=pk, school=request.user.school)
        except Student.DoesNotExist:
            return None

    def post(self, request, pk):
        student = self.get_student(request, pk)
        if not student:
            return ApiResponse.error(message="Student not found.", status_code=404)

        serializer = EnrollStudentSerializer(
            data=request.data,
            context={
                "school": request.user.school,
                "student": student,
            },
        )
        serializer.is_valid(raise_exception=True)

        enrollment = Enrollment.objects.create(
            school=request.user.school,
            student=student,
            klass=serializer.validated_data["klass"],
            academic_year=serializer.validated_data["academic_year"],
        )

        log_action(
            action=AuditLog.Action.CREATE,
            resource="Enrollment",
            resource_id=str(enrollment.pk),
            description=(
                f"{student.full_name} enrolled in "
                f"{enrollment.klass.name} for {enrollment.academic_year.name}"
            ),
            request=request,
        )

        return ApiResponse.created(
            data=EnrollmentSerializer(enrollment).data,
            message="Student enrolled successfully.",
        )

    def delete(self, request, pk):
        """Withdraw a student from their current enrollment."""
        student = self.get_student(request, pk)
        if not student:
            return ApiResponse.error(message="Student not found.", status_code=404)

        academic_year_id = request.data.get("academic_year_id")
        if not academic_year_id:
            return ApiResponse.error(message="academic_year_id is required.")

        try:
            enrollment = Enrollment.objects.get(
                student=student,
                academic_year_id=academic_year_id,
                school=request.user.school,
                is_active=True,
            )
        except Enrollment.DoesNotExist:
            return ApiResponse.error(
                message="Active enrollment not found.", status_code=404
            )

        enrollment.is_active = False
        enrollment.save()

        log_action(
            action=AuditLog.Action.UPDATE,
            resource="Enrollment",
            resource_id=str(enrollment.pk),
            description=f"{student.full_name} withdrawn from {enrollment.klass.name}",
            request=request,
        )

        return ApiResponse.success(message="Student withdrawn from class successfully.")


class StudentEnrollmentHistoryView(generics.ListAPIView):
    """Full enrollment history for a student."""
    permission_classes = [IsAdminOrTeacher]
    serializer_class = EnrollmentSerializer

    def get_queryset(self):
        return Enrollment.objects.filter(
            student_id=self.kwargs["pk"],
            school=self.request.user.school,
        ).select_related("klass", "academic_year")

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        return ApiResponse.success(data=serializer.data)


class ClassStudentsListView(generics.ListAPIView):
    '''List of students currently enrolled in a specific class.'''
    permission_classes = [IsAdminOrTeacher]
    serializer_class = StudentMinimalSerializer

    def get_queryset(self):
        return Student.objects.filter(
            school=self.request.user.school,
            enrollments__klass_id=self.kwargs["class_id"],
            enrollments__is_active=True,
        ).distinct().prefetch_related("guardians","enrollments__klass", "enrollments__academic_year")
    
    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True)
        return ApiResponse.success(data=serializer.data)